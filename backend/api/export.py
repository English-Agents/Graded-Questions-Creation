"""
Excel export endpoint.
GET /api/export/questions  — export question pool to .xlsx

Column layout:
  Course | Module | Topic | Skill | Marks | Question Type | Difficulty |
  Bloom Level | Question Text | Answer Key | Option A | Option B | Option C | Option D |
  Validation Status
"""
from io import BytesIO

import openpyxl
from fastapi import APIRouter, Depends, Response
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from backend.db.connection import get_db
from backend.db.models import Course, Module, PoolConfig, Question, Skill, Topic

router = APIRouter(prefix="/api/export", tags=["export"])

HEADERS = [
    "Course", "Module", "Topic",
    "Skill", "Marks", "Question Type", "Difficulty", "Bloom Level",
    "Question Text", "Answer Key",
    "Option A", "Option B", "Option C", "Option D",
    "Validation Status",
]

BLOOM_LABELS = {1: "K1 Remember", 2: "K2 Understand", 3: "K3 Apply",
                4: "K4 Analyse", 5: "K5 Evaluate", 6: "K6 Create"}

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
APPROVED_FILL = PatternFill("solid", fgColor="E8F5E9")
REJECTED_FILL = PatternFill("solid", fgColor="FFEBEE")
WRAP = Alignment(wrap_text=True, vertical="top")


def _style_header(ws):
    for col_num, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP

    # Column widths (approximate)
    col_widths = [18, 30, 35, 22, 8, 30, 10, 14, 80, 60, 25, 25, 25, 25, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


@router.get("/questions")
async def export_questions(
    topic_id: int | None = None,
    skill: str | None = None,
    difficulty: str | None = None,
    status: str | None = None,
    db: AsyncSession = Depends(get_db),
):
    """
    Export matching questions to an Excel file.
    All filter parameters are optional; without filters all approved questions are exported.
    """
    # Default: export only approved questions unless caller overrides
    if status is None:
        status = "approved"

    stmt = (
        select(Question, PoolConfig, Skill, Topic, Module, Course)
        .join(PoolConfig, PoolConfig.id == Question.pool_config_id)
        .join(Skill, Skill.id == PoolConfig.skill_id)
        .join(Topic, Topic.id == PoolConfig.topic_id)
        .join(Module, Module.id == Topic.module_id)
        .join(Course, Course.id == Module.course_id)
    )

    if topic_id is not None:
        stmt = stmt.where(PoolConfig.topic_id == topic_id)
    if skill:
        stmt = stmt.where(Skill.name == skill)
    if difficulty:
        stmt = stmt.where(Question.difficulty == difficulty)
    if status:
        stmt = stmt.where(Question.validation_status == status)

    stmt = stmt.order_by(Course.name, Module.order_index, Topic.id, Question.created_at)
    result = await db.execute(stmt)
    rows = result.all()

    # Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Questions"
    ws.freeze_panes = "A2"

    _style_header(ws)

    for row_num, (q, pc, sk, topic, mod, course) in enumerate(rows, start=2):
        opts = q.options or {}
        bloom_label = BLOOM_LABELS.get(q.bloom_level, "") if q.bloom_level else ""

        values = [
            course.name,
            mod.name,
            topic.name,
            sk.name,
            pc.marks,
            pc.question_type,
            q.difficulty,
            bloom_label,
            q.text,
            q.answer_key or "",
            opts.get("a", ""),
            opts.get("b", ""),
            opts.get("c", ""),
            opts.get("d", ""),
            q.validation_status,
        ]

        row_fill = None
        if q.validation_status == "approved":
            row_fill = APPROVED_FILL
        elif q.validation_status == "rejected":
            row_fill = REJECTED_FILL

        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = WRAP
            if row_fill:
                cell.fill = row_fill

    # Auto-filter
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(HEADERS))}1"

    # Save to bytes
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = "questions_export.xlsx"
    if topic_id:
        filename = f"questions_topic_{topic_id}.xlsx"

    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
