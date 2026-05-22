"""
app.py — GA Assessment Question Generator (Streamlit Frontend)

Run from the project root:
    streamlit run app.py
"""

import io
import json
import os
import re
import sys
from pathlib import Path
from typing import Optional

import openpyxl
import streamlit as st
import yaml
from openpyxl.styles import Alignment, Font, PatternFill

# ── Paths ────────────────────────────────────────────────────────────────────
ROOT = Path(__file__).parent
COURSES_DIR = ROOT / "courses"
EVALS_DIR = ROOT / "evals"

# ── Question types ───────────────────────────────────────────────────────────
QUESTION_TYPES = [
    "Cloze",
    "Fill in the Blanks",
    "Error Correction",
    "Sentence Arrangement",
    "Jumbled Sentences",
    "Sentence Conversion",
    "Reading Comprehension (higher-order)",
    "Reading Comprehension (literal and inferential)",
    "Reading Comprehension (choice-based)",
    "Short functional writing",
    "Essay writing",
    "Story writing",
    "Email writing",
    "Notice writing",
    "Report / Work report",
    "Process writing / sequencing",
    "Paragraph writing (choice-based)",
    "Story writing (long)",
    "Sequencing",
]

BLOOM_LEVELS = ["L1", "L2", "L3", "L4", "L5", "L6"]
BLOOM_LABELS = {
    "L1": "L1 — Remember",
    "L2": "L2 — Understand",
    "L3": "L3 — Apply",
    "L4": "L4 — Analyse",
    "L5": "L5 — Evaluate",
    "L6": "L6 — Create",
}

DIFFICULTIES = ["Easy", "Medium", "Hard", "Medium to Hard"]

FOLDER_MAP = {
    "Cloze":                "cloze",
    "Fill in the Blanks":   "fill_in_the_blanks",
    "Error Correction":     "error_correction",
    "Sentence Arrangement": "sentence_arrangement",
    "Jumbled Sentences":    "jumbled_sentences",
    "Sentence Conversion":  "sentence_conversion",
}

TYPE_RULES = {
    "Cloze": """
Each question must be a passage with EXACTLY 4 numbered blanks (i)–(iv).
Each blank offers TWO bracket options: one noun form and one verb form.
Format example: the manager (i)_________________ (approval / approved) the report.
The student chooses the correct word from the two options in brackets.
""",
    "Fill in the Blanks": """
Each question must have EXACTLY 2 numbered blanks (i)–(ii).
Each blank shows ONE base word in brackets — the student derives the correct noun or verb form.
Format example: the director gave a clear (i)_________________ (explain) of the plan.
One blank must require a noun form; the other a main verb form.
""",
    "Error Correction": """
Each question must be a single sentence containing EXACTLY 2 incorrect words:
one noun used where a verb is needed, and one verb used where a noun is needed.
The instruction must say: "Identify the incorrect words and replace them with the correct forms."
Include the corrected sentence in the solution.
""",
    "Sentence Arrangement": """
Each question presents 4–5 parts of a sentence labelled (A), (B), (C), (D)[, (E)].
The student arranges them into one coherent sentence.
The solution gives the correct order (e.g., "D–B–A–C").
""",
    "Jumbled Sentences": """
Each question presents 4–6 jumbled sentences that form a coherent paragraph.
The student arranges them in the correct order.
The solution gives the correct sequence.
""",
    "Sentence Conversion": """
Each question gives one complete sentence and asks the student to rewrite it using a
specified grammatical structure (e.g. active → passive, direct → indirect speech).
The instruction must name the required transformation clearly.
""",
}

DEFAULT_TYPE_RULE = """
Generate well-formed questions matching the specified type.
Include clear instructions, question text, solution, and explanation.
"""

COURSE_DISPLAY = {
    "foundation":        "English Foundation",
    "advanced":          "English Advanced",
    "applied":           "English Applied",
    "language_analytics": "English Language Analytics",
}


# ── Discovery helpers ─────────────────────────────────────────────────────────

def get_courses() -> list[str]:
    if not COURSES_DIR.exists():
        return []
    return sorted(d.name for d in COURSES_DIR.iterdir() if d.is_dir())


def get_modules(course: str) -> list[str]:
    course_dir = COURSES_DIR / course
    return sorted(
        d.name for d in course_dir.iterdir()
        if d.is_dir() and d.name.startswith("module_")
    )


def get_topics(course: str, module: str) -> list[str]:
    module_dir = COURSES_DIR / course / module
    return sorted(
        d.name for d in module_dir.iterdir()
        if d.is_dir() and d.name.startswith("topic_")
    )


def get_material_file(course: str, module: str, topic: str) -> Optional[Path]:
    mat_dir = COURSES_DIR / course / module / topic / "material"
    if not mat_dir.exists():
        return None
    files = sorted(mat_dir.glob("*.md"))
    return files[0] if files else None


def load_metadata(course: str, module: str, topic: str) -> dict:
    meta_file = COURSES_DIR / course / module / topic / "metadata.json"
    if meta_file.exists():
        try:
            return json.loads(meta_file.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def topic_display_name(course: str, module: str, topic: str) -> str:
    meta = load_metadata(course, module, topic)
    if meta.get("topic_name"):
        return meta["topic_name"]
    mat = get_material_file(course, module, topic)
    if mat:
        return mat.stem.replace("_", " ").title()
    return topic.replace("_", " ").title()


def module_display_name(course: str, module: str) -> str:
    topic_list = get_topics(course, module)
    if topic_list:
        meta = load_metadata(course, module, topic_list[0])
        if meta.get("module"):
            return meta["module"]
    num = module.split("_")[-1]
    return f"Module {num}"


# ── Content helpers ───────────────────────────────────────────────────────────

def load_material(path: Path) -> str:
    text = path.read_text(encoding="utf-8")
    text = re.sub(r"<[^>]+>", " ", text)          # strip XML/HTML tags
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def load_samples(question_type: str) -> list[dict]:
    folder = FOLDER_MAP.get(question_type)
    if not folder:
        return []
    eval_file = EVALS_DIR / folder / "eval_tests.yaml"
    if not eval_file.exists():
        return []
    raw = yaml.safe_load(eval_file.read_text(encoding="utf-8"))
    if not raw:
        return []
    return [item.get("vars", {}) for item in raw if isinstance(item, dict)]


def format_samples(samples: list[dict]) -> str:
    if not samples:
        return "No sample questions available for this type yet."
    lines = []
    for i, s in enumerate(samples[:3], 1):
        lines.append(f"--- SAMPLE {i} ---")
        lines.append(f"Question:\n{s.get('question', '').strip()}")
        lines.append(f"\nSolution:\n{s.get('solution', '').strip()}")
        lines.append(f"\nExplanation:\n{s.get('explanation', '').strip()}")
        lines.append("")
    return "\n".join(lines)


# ── Generation ────────────────────────────────────────────────────────────────

def build_prompt(material: str, question_type: str, count: int,
                 bloom: str, difficulty: str, course_outcome: str,
                 samples: list[dict]) -> str:
    rules = TYPE_RULES.get(question_type, DEFAULT_TYPE_RULE).strip()
    sample_text = format_samples(samples)

    return f"""You are an expert English language assessment designer specialising in {question_type} questions.

══════════════════════════════════════════
READING MATERIAL (use as source for all questions)
══════════════════════════════════════════
{material[:7000]}

══════════════════════════════════════════
QUESTION TYPE RULES — {question_type.upper()}
══════════════════════════════════════════
{rules}

══════════════════════════════════════════
SAMPLE QUESTIONS (follow this format exactly)
══════════════════════════════════════════
{sample_text}

══════════════════════════════════════════
YOUR TASK
══════════════════════════════════════════
Generate {count} new {question_type} question(s).

Parameters:
  Bloom's Level  : {bloom}
  Difficulty     : {difficulty}
  Course Outcome : {course_outcome}

Rules:
- All scenarios must come from the reading material above.
- Do NOT copy the sample questions — create entirely new scenarios.
- Follow the exact format of the samples for question, solution, and explanation.
- Explanations must use grammatical reasoning (e.g. "follows the adjective", "main verb showing the action").

Use this EXACT output structure for each question — no deviations:

========================================
QUESTION [n]
========================================
Question:
<full question text including instruction line>

Solution:
<correct answer>

Explanation:
<grammatical justification>
"""


def call_anthropic(prompt: str, model: str, api_key: str) -> str:
    from anthropic import Anthropic
    client = Anthropic(api_key=api_key)
    response = client.messages.create(
        model=model,
        max_tokens=4096,
        messages=[{"role": "user", "content": prompt}],
    )
    return response.content[0].text


# ── Output parsing ────────────────────────────────────────────────────────────

def parse_questions(raw: str) -> list[dict]:
    blocks = re.split(r"={3,}\s*QUESTION\s+\d+\s*={3,}", raw, flags=re.IGNORECASE)
    results = []
    for block in blocks:
        block = block.strip()
        if not block:
            continue

        def extract(label: str) -> str:
            pattern = rf"{label}:\s*\n(.*?)(?=\n(?:Question|Solution|Explanation):|$)"
            m = re.search(pattern, block, re.DOTALL | re.IGNORECASE)
            return m.group(1).strip() if m else ""

        q = extract("Question")
        s = extract("Solution")
        e = extract("Explanation")
        if q:
            results.append({"question": q, "solution": s, "explanation": e})
    return results


# ── Excel export ──────────────────────────────────────────────────────────────

def build_excel(questions: list[dict], meta: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Generated Questions"

    # Header
    headers = [
        "Question no.", "Module Number", "Blooms Level", "Difficulty level",
        "Course Outcomes", "Name of the Module", "Topic name", "Questions", "Solution",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30

    # Data rows — alternating colours
    fill_a = PatternFill("solid", fgColor="DEEAF1")
    fill_b = PatternFill("solid", fgColor="FFFFFF")

    module_num = meta.get("module", "").split("_")[-1] if "_" in meta.get("module", "") else meta.get("module", "")

    for i, q in enumerate(questions, 1):
        row = i + 1
        values = [
            i,
            module_num,
            meta.get("bloom", ""),
            meta.get("difficulty", ""),
            meta.get("course_outcome", ""),
            meta.get("module_display", ""),
            meta.get("topic_display", ""),
            q.get("question", ""),
            q.get("solution", ""),
        ]
        row_fill = fill_a if i % 2 == 0 else fill_b
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = row_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Column widths
    widths = [12, 14, 12, 14, 40, 28, 28, 60, 40]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Streamlit App ─────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Assessment Question Generator",
    page_icon="📚",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.image("https://img.icons8.com/color/96/book.png", width=60)
    st.title("📚 Question Generator")
    st.markdown("---")

    # API Key
    st.subheader("🔑 API Key")
    env_key = ""
    env_file = ROOT / ".env"
    env_example = ROOT / ".env.example"
    for f in [env_file, env_example]:
        if f.exists():
            for line in f.read_text().splitlines():
                if line.startswith("ANTHROPIC_API_KEY="):
                    env_key = line.split("=", 1)[1].strip()
                    break
        if env_key:
            break

    api_key = st.text_input(
        "Anthropic API Key",
        value=env_key,
        type="password",
        help="Get your key from console.anthropic.com",
    )
    if api_key:
        st.success("✅ API key loaded")
    else:
        st.warning("⚠️ Enter your API key above")

    st.markdown("---")

    # Model
    st.subheader("🤖 Model")
    model = st.selectbox(
        "Claude Model",
        options=[
            "claude-haiku-4-5-20251001",
            "claude-sonnet-4-6",
        ],
        index=0,
        help="Haiku is faster and cheaper. Sonnet gives better quality.",
    )

    st.markdown("---")

    # Course → Module → Topic
    st.subheader("📁 Select Topic")
    courses = get_courses()
    if not courses:
        st.error("No courses found. Check the `courses/` folder.")
        st.stop()

    selected_course = st.selectbox(
        "Course",
        courses,
        format_func=lambda c: COURSE_DISPLAY.get(c, c.replace("_", " ").title()),
    )

    modules = get_modules(selected_course)
    if not modules:
        st.warning("No modules found in this course.")
        st.stop()

    selected_module = st.selectbox(
        "Module",
        modules,
        format_func=lambda m: module_display_name(selected_course, m),
    )

    topics = get_topics(selected_course, selected_module)
    if not topics:
        st.warning("No topics found in this module.")
        st.stop()

    selected_topic = st.selectbox(
        "Topic",
        topics,
        format_func=lambda t: topic_display_name(selected_course, selected_module, t),
    )

    material_file = get_material_file(selected_course, selected_module, selected_topic)
    if material_file:
        st.success(f"✅ Material: `{material_file.name}`")
        mat_size = material_file.stat().st_size
        st.caption(f"{mat_size // 1024} KB of reading material available")
    else:
        st.error("❌ No material file found for this topic. Add a .md file to the material/ folder.")


# ── Main area ─────────────────────────────────────────────────────────────────

topic_display = topic_display_name(selected_course, selected_module, selected_topic)
module_display = module_display_name(selected_course, selected_module)
course_display = COURSE_DISPLAY.get(selected_course, selected_course.replace("_", " ").title())

st.header(f"Generate Questions — {course_display}")
st.markdown(f"**{module_display}** › **{topic_display}**")
st.markdown("---")

col1, col2, col3 = st.columns(3)

with col1:
    question_type = st.selectbox("Question Type", QUESTION_TYPES)
    count = st.slider("Number of Questions", min_value=1, max_value=10, value=5)

with col2:
    bloom_label = st.selectbox(
        "Bloom's Level",
        options=list(BLOOM_LABELS.values()),
        index=3,
    )
    bloom = bloom_label.split(" ")[0]
    difficulty = st.selectbox("Difficulty", DIFFICULTIES, index=1)

with col3:
    meta_data = load_metadata(selected_course, selected_module, selected_topic)
    default_co = ""
    if meta_data.get("skills"):
        default_co = "CO1: "
    course_outcome = st.text_area(
        "Course Outcome",
        value=default_co,
        height=100,
        placeholder="CO1: Identify and correct errors in the use of nouns and main verbs...",
        help="Describe what the student should be able to do after studying this topic.",
    )

st.markdown("---")

# Generate button
can_generate = bool(api_key) and material_file is not None
if not can_generate:
    if not api_key:
        st.info("👈 Enter your Anthropic API key in the sidebar to get started.")
    if material_file is None:
        st.warning("This topic has no reading material yet. Add a `.md` file to the material folder.")

generate_clicked = st.button(
    f"🚀 Generate {count} Question(s)",
    type="primary",
    disabled=not can_generate,
    use_container_width=True,
)

if generate_clicked:
    with st.spinner(f"Generating {count} × {question_type} question(s)... this takes ~20–40 seconds"):
        try:
            material_text = load_material(material_file)
            samples = load_samples(question_type)
            prompt = build_prompt(
                material=material_text,
                question_type=question_type,
                count=count,
                bloom=bloom,
                difficulty=difficulty,
                course_outcome=course_outcome,
                samples=samples,
            )
            raw_output = call_anthropic(prompt, model, api_key)

            parsed = parse_questions(raw_output)

            st.session_state["raw_output"] = raw_output
            st.session_state["parsed"] = parsed
            st.session_state["gen_meta"] = {
                "course":          selected_course,
                "course_display":  course_display,
                "module":          selected_module,
                "module_display":  module_display,
                "topic":           selected_topic,
                "topic_display":   topic_display,
                "question_type":   question_type,
                "bloom":           bloom,
                "difficulty":      difficulty,
                "course_outcome":  course_outcome,
            }

        except Exception as e:
            err = str(e)
            if "credit balance" in err.lower():
                st.error("❌ Your Anthropic API credit balance is too low. Please top up at console.anthropic.com/settings/billing")
            elif "invalid api key" in err.lower() or "authentication" in err.lower():
                st.error("❌ Invalid API key. Check the key in the sidebar.")
            else:
                st.error(f"❌ Generation failed: {err}")

# ── Results ───────────────────────────────────────────────────────────────────

if "parsed" in st.session_state and st.session_state["parsed"]:
    parsed: list[dict] = st.session_state["parsed"]
    meta: dict = st.session_state["gen_meta"]

    st.markdown("---")
    st.subheader(f"✅ {len(parsed)} Question(s) Generated")

    # Display each question
    for i, q in enumerate(parsed, 1):
        with st.expander(f"Question {i}", expanded=True):
            st.markdown("**Question:**")
            st.markdown(q.get("question", ""))
            st.markdown("**Solution:**")
            st.code(q.get("solution", ""), language=None)
            st.markdown("**Explanation:**")
            st.markdown(q.get("explanation", ""))

    st.markdown("---")

    # Download row
    col_dl1, col_dl2 = st.columns(2)

    with col_dl1:
        excel_bytes = build_excel(parsed, meta)
        filename = f"{meta['course']}_{meta['module']}_{meta['topic']}_{meta['question_type'].replace(' ', '_')}.xlsx"
        st.download_button(
            label="⬇️ Download as Excel",
            data=excel_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    with col_dl2:
        st.download_button(
            label="⬇️ Download raw text",
            data=st.session_state["raw_output"],
            file_name=filename.replace(".xlsx", ".txt"),
            mime="text/plain",
            use_container_width=True,
        )

    # Summary strip
    st.markdown("---")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Questions", len(parsed))
    c2.metric("Type", meta["question_type"])
    c3.metric("Bloom's", meta["bloom"])
    c4.metric("Difficulty", meta["difficulty"])

elif "raw_output" in st.session_state and not st.session_state.get("parsed"):
    st.warning("Questions were generated but could not be parsed into structured form. Showing raw output:")
    st.code(st.session_state["raw_output"])
